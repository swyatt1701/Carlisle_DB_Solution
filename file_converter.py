import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from datetime import datetime
import os
import threading

class FileConverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("File Converter")
        self.root.geometry("500x500")  # Increased height to accommodate changes
        self.root.resizable(False, False)
        
        # Configure style
        self.setup_styles()
        
        # Variables
        self.source_file = tk.StringVar()
        self.dest_file = tk.StringVar()
        
        # Create GUI elements
        self.create_widgets()
        
    def setup_styles(self):
        """Configure the visual styling"""
        self.root.configure(bg='#f0f0f0')
        
    def create_widgets(self):
        """Create all GUI widgets"""
        
        # Main frame with padding
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(
            main_frame, 
            text="File Converter",
            font=('Arial', 16, 'bold'),
            bg='#f0f0f0'
        )
        title_label.pack(pady=(0, 30))
        
        # Select File section
        self.create_file_selection_section(main_frame)
        
        # Select Destination section  
        self.create_destination_section(main_frame)
        
        # Convert button
        self.create_convert_section(main_frame)
        
        # Status/Progress section
        self.create_status_section(main_frame)
        
    def create_file_selection_section(self, parent):
        """Create the file selection section"""
        
        # Select File frame
        select_frame = tk.Frame(parent, bg='#f0f0f0')
        select_frame.pack(fill='x', pady=(0, 20))
        
        # Select File button
        select_btn = tk.Button(
            select_frame,
            text="Select File",
            font=('Arial', 12),
            bg='#e1e1e1',
            relief='raised',
            borderwidth=2,
            command=self.select_source_file,
            width=15,
            height=2
        )
        select_btn.pack()
        
        # File path display - doubled height with scrollbar
        path_frame = tk.Frame(select_frame, bg='white', relief='sunken', borderwidth=1, height=80)
        path_frame.pack(fill='x', pady=(10, 0))
        path_frame.pack_propagate(False)  # Do not allow resizing
        
        # Text widget with scrollbar
        self.source_text = tk.Text(
            path_frame,
            height=4,  # Doubled from 2 to 4
            bg='white',
            fg='red',
            font=('Arial', 9),
            wrap='word',
            state='disabled'
        )
        
        # Scrollbar for source text
        source_scrollbar = tk.Scrollbar(path_frame, command=self.source_text.yview)
        self.source_text.configure(yscrollcommand=source_scrollbar.set)
        
        self.source_text.pack(side='left', fill='both', expand=True, padx=5, pady=5)
        source_scrollbar.pack(side='right', fill='y')
        
    def create_destination_section(self, parent):
        """Create the destination selection section"""
        
        # Select Destination frame
        dest_frame = tk.Frame(parent, bg='#f0f0f0')
        dest_frame.pack(fill='x', pady=(0, 20))
        
        # Select Destination button
        dest_btn = tk.Button(
            dest_frame,
            text="Select Destination",
            font=('Arial', 12),
            bg='#e1e1e1',
            relief='raised',
            borderwidth=2,
            command=self.select_destination,
            width=15,
            height=2
        )
        dest_btn.pack()
        
        # Destination path display - with scrollbar, back to 2 lines
        dest_path_frame = tk.Frame(dest_frame, bg='white', relief='sunken', borderwidth=1, height=40)
        dest_path_frame.pack(fill='x', pady=(10, 0))
        dest_path_frame.pack_propagate(False)  # Do not allow resizing
        
        # Text widget with scrollbar
        self.dest_text = tk.Text(
            dest_path_frame,
            height=2,  # Back to 2 lines
            bg='white',
            fg='red',
            font=('Arial', 9),
            wrap='word',
            state='disabled'
        )
        
        # Scrollbar for dest text
        dest_scrollbar = tk.Scrollbar(dest_path_frame, command=self.dest_text.yview)
        self.dest_text.configure(yscrollcommand=dest_scrollbar.set)
        
        self.dest_text.pack(side='left', fill='both', expand=True, padx=5, pady=5)
        dest_scrollbar.pack(side='right', fill='y')
        
        # Note about EXPORT filename
        note_label = tk.Label(
            dest_frame,
            text='Note: Save the file as "EXPORT"',
            font=('Arial', 9, 'italic'),
            bg='#f0f0f0',
            fg='#666666'
        )
        note_label.pack(pady=(5, 0), anchor='w')
        
    def create_convert_section(self, parent):
        """Create the convert button section"""
        
        # Convert button
        self.convert_btn = tk.Button(
            parent,
            text="Convert File",
            font=('Arial', 14, 'bold'),
            bg='#4CAF50',
            fg='white',
            relief='raised',
            borderwidth=3,
            command=self.on_convert_click,
            width=12,
            height=2
        )
        self.convert_btn.pack(pady=20)
        
    def create_status_section(self, parent):
        """Create the status display section"""
        
        # Status text area
        self.status_text = tk.Text(
            parent,
            height=8,
            width=60,
            font=('Consolas', 9),
            bg='#f8f8f8',
            relief='sunken',
            borderwidth=2,
            state='disabled'
        )
        self.status_text.pack(fill='both', expand=True, pady=(10, 0))
        
        # Add scrollbar
        scrollbar = tk.Scrollbar(self.status_text)
        scrollbar.pack(side='right', fill='y')
        self.status_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.status_text.yview)
        
    def select_source_file(self):
        """Handle source file selection"""
        filename = filedialog.askopenfilename(
            title="Select CenterPoint Export File",
            filetypes=[
                ("Excel files", "*.xls *.xlsx"),
                ("All files", "*.*")
            ]
        )
        
        if filename:
            self.source_file.set(filename)
            # Update the text widget
            self.source_text.config(state='normal')
            self.source_text.delete('1.0', 'end')
            self.source_text.insert('1.0', filename)
            self.source_text.config(state='disabled')
            self.log_status(f"Selected source file: {os.path.basename(filename)}")
            
    def select_destination(self):
        """Handle destination selection"""
        filename = filedialog.asksaveasfilename(
            title="Save Converted File As",
            defaultextension=".xls",
            filetypes=[
                ("Excel 97-2003", "*.xls"),
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ]
        )
        
        if filename:
            self.dest_file.set(filename)
            # Update the text widget
            self.dest_text.config(state='normal')
            self.dest_text.delete('1.0', 'end')
            self.dest_text.insert('1.0', filename)
            self.dest_text.config(state='disabled')
            self.log_status(f"Selected destination: {os.path.basename(filename)}")
    
    def log_status(self, message):
        """Add message to status log"""
        self.status_text.config(state='normal')
        self.status_text.insert('end', f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.status_text.see('end')
        self.status_text.config(state='disabled')
        self.root.update()
        
    def convert_file(self):
        """Handle file conversion in separate thread"""
        
        if not self.source_file.get():
            messagebox.showerror("Error", "Please select a source file first")
            return
            
        if not self.dest_file.get():
            messagebox.showerror("Error", "Please select a destination first")
            return
        
        # Disable convert button during processing
        self.convert_btn.config(state='disabled')
        
        # Run conversion in separate thread to prevent GUI freezing
        thread = threading.Thread(target=self.perform_conversion)
        thread.daemon = True
        thread.start()
        
    def perform_conversion(self):
        """Perform the actual file conversion"""
        try:
            source = self.source_file.get()
            destination = self.dest_file.get()
            
            # Read the CenterPoint file
            cp_df = pd.read_excel(source, header=None)
            
            # Create new workbook matching TurningPoint structure
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Check Register"
            
            # Set up TurningPoint header structure
            today_excel_date = (datetime.now() - datetime(1900, 1, 1)).days + 2
            ws['A1'] = today_excel_date
            ws['F1'] = "Carlisle County Fiscal Court"
            ws['K1'] = "Page -1 of 1"
            
            ws['F3'] = "Check Register"
            ws['F4'] = "Checks with Account Detail"
            
            # Column headers
            headers_row7 = ["Check", None, "Check", "Bank", "Vendor", "Vendor", "Invoice", "Invoice", "Invoice", None, "Check"]
            headers_row8 = ["Number", None, "Date", "Code", "Code", "Description", "Number", "Date", "Amount", None, "Amount"]
            
            for col, header in enumerate(headers_row7, 1):
                if header:
                    ws.cell(row=7, column=col, value=header)
            
            for col, header in enumerate(headers_row8, 1):
                if header:
                    ws.cell(row=8, column=col, value=header)
            
            # Find data start row
            data_start_row = None
            for idx, row in cp_df.iterrows():
                if pd.notna(row.iloc[1]) and pd.notna(row.iloc[3]):
                    try:
                        # Check if this looks like a data row (has check number)
                        check_num = str(row.iloc[3]).strip()
                        if check_num and not "Check" in check_num and not "Number" in check_num:
                            data_start_row = idx
                            break
                    except:
                        continue
            
            if data_start_row is None:
                raise Exception("Could not find data start row in CenterPoint file")
            
            current_output_row = 11
            records_converted = 0
            current_date = None  # Track the current date to carry forward
            warnings = []  # Track missing data warnings
            
            for idx in range(data_start_row, len(cp_df)):
                cp_row = cp_df.iloc[idx]
                
                # Skip empty rows or summary rows
                if pd.isna(cp_row.iloc[3]) and pd.isna(cp_row.iloc[4]) and pd.isna(cp_row.iloc[6]):
                    continue
                
                # Extract CenterPoint data
                row_date = cp_row.iloc[1] if pd.notna(cp_row.iloc[1]) else None
                check_number = str(cp_row.iloc[3]).strip() if pd.notna(cp_row.iloc[3]) and str(cp_row.iloc[3]).strip() else ""
                account_number = str(cp_row.iloc[4]).strip() if pd.notna(cp_row.iloc[4]) and str(cp_row.iloc[4]).strip() else ""
                amount = cp_row.iloc[5] if pd.notna(cp_row.iloc[5]) else 0
                vendor = str(cp_row.iloc[6]).strip() if pd.notna(cp_row.iloc[6]) and str(cp_row.iloc[6]).strip() else ""
                
                # Skip header rows or invalid data
                if "Check" in check_number or "Number" in check_number:
                    continue
                
                # Check for missing critical data and create warnings
                row_warnings = []
                if not check_number:
                    row_warnings.append("missing check number")
                if not account_number:
                    row_warnings.append("missing account number")
                if not vendor:
                    row_warnings.append("missing vendor name")
                
                # If we have warnings, record them with transaction details
                if row_warnings:
                    warning_msg = f"Row {idx + 1}: Transaction imported with {', '.join(row_warnings)}"
                    if amount:
                        warning_msg += f" (Amount: ${amount})"
                    if check_number:
                        warning_msg += f" (Check: {check_number})"
                    warnings.append(warning_msg)
                
                # Update current_date if this row has a date, otherwise use the last known date
                if row_date is not None:
                    current_date = row_date
                
                # Use current_date for this transaction (never null)
                transaction_date = current_date if current_date is not None else ""
                
                # Map to TurningPoint format (import even with missing data)
                ws.cell(row=current_output_row, column=1, value=check_number if check_number else "")
                ws.cell(row=current_output_row, column=3, value=transaction_date)
                ws.cell(row=current_output_row, column=4, value="General")
                ws.cell(row=current_output_row, column=5, value=vendor[:20] if vendor else "")
                ws.cell(row=current_output_row, column=6, value=vendor if vendor else "")
                ws.cell(row=current_output_row, column=7, value="")
                ws.cell(row=current_output_row, column=8, value=transaction_date)
                ws.cell(row=current_output_row, column=9, value=amount)
                ws.cell(row=current_output_row, column=11, value=amount)
                
                # Add account detail row
                current_output_row += 1
                ws.cell(row=current_output_row, column=3, value="Account:")
                ws.cell(row=current_output_row, column=4, value=account_number if account_number else "")
                ws.cell(row=current_output_row, column=5, value="Amount:")
                ws.cell(row=current_output_row, column=6, value=amount)
                
                current_output_row += 2
                records_converted += 1
            
            # Save the file
            wb.save(destination)
            
            # Prepare success message with warnings
            success_message = f"File converted successfully!\n\nRecords converted: {records_converted}\nFile saved as: {os.path.basename(destination)}\n\nThe file is ready to import into Access!"
            
            if warnings:
                warning_text = "\n\nWARNINGS - Manual data entry required:\n" + "\n".join(warnings)
                success_message += warning_text
                
                # Show warnings in a separate dialog if there are many
                if len(warnings) > 5:
                    messagebox.showwarning(
                        "Data Quality Warnings", 
                        f"Conversion completed with {len(warnings)} warnings:\n\n" + 
                        "\n".join(warnings[:5]) + 
                        f"\n\n... and {len(warnings) - 5} more warnings.\n\n" +
                        "Please review the converted file and manually enter missing data."
                    )
            
            messagebox.showinfo("Conversion Complete", success_message)
            
        except Exception as e:
            error_msg = f"Conversion failed: {str(e)}"
            messagebox.showerror("Conversion Error", error_msg)
            
        finally:
            # Re-enable convert button
            self.convert_btn.config(state='normal')
    
    def on_convert_click(self):
        """Handle convert button click"""
        if not self.source_file.get():
            messagebox.showerror("Error", "Please select a source file first")
            return
            
        if not self.dest_file.get():
            messagebox.showerror("Error", "Please select a destination first")
            return
        
        # Confirm before conversion
        response = messagebox.askyesno(
            "Confirm Conversion", 
            f"Convert CenterPoint file to Access format?\n\n"
            f"Source: {os.path.basename(self.source_file.get())}\n"
            f"Destination: {os.path.basename(self.dest_file.get())}"
        )
        
        if response:
            self.convert_file()

def main():
    """Main application entry point"""
    root = tk.Tk()
    app = FileConverterGUI(root)
    
    # Center window on screen
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")
    
    root.mainloop()

if __name__ == "__main__":
    main()